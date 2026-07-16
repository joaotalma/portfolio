# -*- coding: utf-8 -*-
"""
Relatório macroeconômico automatizado em Excel
==============================================
Gera (e atualiza sozinho) um relatório Excel formatado com três indicadores
públicos do Banco Central (API SGS):

  - 4189: Taxa Selic acumulada no mês, anualizada (% a.a.)
  - 3698: Câmbio — dólar americano, média mensal de venda (R$/US$)
  -  433: IPCA — variação mensal (%)

Saídas:
  - relatorio_macro.xlsx  (capa com destaques + 1 aba por indicador + gráficos)
  - data/*.csv            (snapshot dos dados)

Uso:
  python gerar_relatorio.py            # tenta API; se offline, usa CSVs locais
  python gerar_relatorio.py --offline  # força uso dos CSVs locais

Automação: o workflow em .github/workflows/atualizar-relatorio.yml roda este
script todo mês e commita o Excel atualizado — automação de verdade, sem clique.

Autor: João Talma · joaotalmaj@gmail.com
"""

import json
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
DATA_DIR = BASE / "data"
API_URL = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.{codigo}/dados?formato=json&dataInicial=01/01/2011"

INDICADORES = {
    "4189": {
        "nome": "Selic",
        "titulo": "Taxa Selic (% a.a.)",
        "arquivo": "sgs_4189_selic.csv",
        "formato": "0.00",
        "sufixo": "% a.a.",
    },
    "3698": {
        "nome": "Câmbio",
        "titulo": "Dólar — média mensal de venda (R$/US$)",
        "arquivo": "sgs_3698_cambio.csv",
        "formato": "0.0000",
        "sufixo": " R$/US$",
    },
    "433": {
        "nome": "IPCA",
        "titulo": "IPCA — variação mensal (%)",
        "arquivo": "sgs_433_ipca.csv",
        "formato": "0.00",
        "sufixo": "% no mês",
    },
}

# Identidade visual (mesma paleta do site)
LARANJA = "E85D04"
LARANJA_ESCURO = "C2410C"
CREME = "FFF8F0"
GRAFITE = "1F2937"
CINZA_TEXTO = "78716C"
LINHA_CLARA = "F0E2D0"


# ----------------------------------------------------------------------------
# Dados
# ----------------------------------------------------------------------------
def baixar_serie(codigo: str) -> pd.DataFrame:
    import urllib.request

    url = API_URL.format(codigo=codigo)
    with urllib.request.urlopen(url, timeout=30) as resp:
        bruto = json.loads(resp.read().decode("utf-8"))
    df = pd.DataFrame(bruto)
    df["data"] = pd.to_datetime(df["data"], format="%d/%m/%Y").dt.strftime("%Y-%m")
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
    return df.dropna().reset_index(drop=True)


def obter_dados(offline: bool) -> dict:
    dados = {}
    for codigo, cfg in INDICADORES.items():
        if not offline:
            try:
                df = baixar_serie(codigo)
                DATA_DIR.mkdir(exist_ok=True)
                df.to_csv(DATA_DIR / cfg["arquivo"], index=False)
                dados[codigo] = df
                print(f"[API] {cfg['nome']}: {len(df)} observações")
                continue
            except Exception as exc:  # noqa: BLE001
                print(f"[aviso] API indisponível p/ {cfg['nome']} ({exc}); usando CSV local")
        df = pd.read_csv(DATA_DIR / cfg["arquivo"])
        df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
        dados[codigo] = df.dropna().reset_index(drop=True)
        print(f"[CSV] {cfg['nome']}: {len(dados[codigo])} observações")
    return dados


def ipca_12m(df: pd.DataFrame) -> float:
    """IPCA acumulado em 12 meses (produto dos fatores mensais)."""
    fatores = 1 + df["valor"].tail(12) / 100
    return round((fatores.prod() - 1) * 100, 2)


# ----------------------------------------------------------------------------
# Excel
# ----------------------------------------------------------------------------
def estilo_titulo(ws, celula, texto, tamanho=20, cor=GRAFITE):
    ws[celula] = texto
    ws[celula].font = Font(name="Calibri", size=tamanho, bold=True, color=cor)


def montar_capa(wb: Workbook, dados: dict) -> None:
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False
    for col, largura in zip("ABCDEFG", (2, 26, 22, 22, 22, 22, 2)):
        ws.column_dimensions[col].width = largura

    # Faixa laranja no topo
    for col in range(1, 8):
        for lin in (1, 2):
            ws.cell(row=lin, column=col).fill = PatternFill("solid", fgColor=LARANJA)
    estilo_titulo(ws, "B4", "Relatório Macroeconômico — Brasil", 22)
    ws["B5"] = f"Gerado automaticamente em {datetime.now():%d/%m/%Y %H:%M} · Fonte: Banco Central do Brasil (API SGS)"
    ws["B5"].font = Font(size=10, color=CINZA_TEXTO)

    selic = dados["4189"]
    cambio = dados["3698"]
    ipca = dados["433"]
    def br(txt: str) -> str:
        return txt.replace(".", ",")

    destaques = [
        ("Selic", br(f'{selic["valor"].iloc[-1]:.2f}') + "% a.a.", f'em {selic["data"].iloc[-1]}'),
        ("Dólar (média mensal)", "R$ " + br(f'{cambio["valor"].iloc[-1]:.4f}'), f'em {cambio["data"].iloc[-1]}'),
        ("IPCA 12 meses", br(f"{ipca_12m(ipca):.2f}") + "%", f'até {ipca["data"].iloc[-1]}'),
        ("IPCA no mês", br(f'{ipca["valor"].iloc[-1]:.2f}') + "%", f'em {ipca["data"].iloc[-1]}'),
    ]
    borda = Border(*(Side(style="thin", color=LINHA_CLARA),) * 4)
    for i, (nome, valor, ref) in enumerate(destaques):
        col = 2 + i
        c_nome, c_valor, c_ref = (ws.cell(row=r, column=col) for r in (7, 8, 9))
        c_nome.value, c_valor.value, c_ref.value = nome, valor, ref
        c_nome.font = Font(size=10, bold=True, color=CINZA_TEXTO)
        c_valor.font = Font(size=16, bold=True, color=LARANJA_ESCURO)
        c_ref.font = Font(size=9, color=CINZA_TEXTO)
        for c in (c_nome, c_valor, c_ref):
            c.fill = PatternFill("solid", fgColor=CREME)
            c.border = borda
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws["B11"] = "Cada indicador tem sua própria aba, com a série mensal completa desde 2011 e gráfico."
    ws["B11"].font = Font(size=10, italic=True, color=CINZA_TEXTO)

    # Impressão: paisagem, ajustada à largura
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


def montar_aba(wb: Workbook, cfg: dict, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(cfg["nome"])
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14

    estilo_titulo(ws, "A1", cfg["titulo"], 14)
    ws["A2"] = "Fonte: Banco Central do Brasil — API SGS"
    ws["A2"].font = Font(size=9, color=CINZA_TEXTO)

    # Cabeçalho da tabela
    for col, txt in (("A", "Mês"), ("B", "Valor")):
        c = ws[f"{col}4"]
        c.value = txt
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=LARANJA)
        c.alignment = Alignment(horizontal="center")

    zebra = PatternFill("solid", fgColor=CREME)
    for i, linha in enumerate(df.itertuples(index=False)):
        r = 5 + i
        ws.cell(row=r, column=1, value=linha.data).alignment = Alignment(horizontal="center")
        c = ws.cell(row=r, column=2, value=float(linha.valor))
        c.number_format = cfg["formato"]
        if i % 2:
            ws.cell(row=r, column=1).fill = zebra
            c.fill = zebra

    # Gráfico de linha
    chart = LineChart()
    chart.title = cfg["titulo"]
    chart.style = 2
    chart.height = 9
    chart.width = 22
    chart.y_axis.majorGridlines = None
    valores = Reference(ws, min_col=2, min_row=4, max_row=4 + len(df))
    categorias = Reference(ws, min_col=1, min_row=5, max_row=4 + len(df))
    chart.add_data(valores, titles_from_data=True)
    chart.set_categories(categorias)
    serie = chart.series[0]
    serie.graphicalProperties.line.solidFill = LARANJA
    serie.graphicalProperties.line.width = 22000
    chart.legend = None
    ws.add_chart(chart, "D4")


def main() -> None:
    offline = "--offline" in sys.argv
    dados = obter_dados(offline)

    wb = Workbook()
    montar_capa(wb, dados)
    for codigo, cfg in INDICADORES.items():
        montar_aba(wb, cfg, dados[codigo])

    destino = BASE / "relatorio_macro.xlsx"
    wb.save(destino)
    print(f"\nRelatorio salvo em {destino}")


if __name__ == "__main__":
    main()
